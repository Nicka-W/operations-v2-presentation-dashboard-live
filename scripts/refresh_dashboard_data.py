"""
refresh_dashboard_data.py

Reads the source workbook's `DATA` sheet and builds dashboard_data.json
for the Operations V2 dashboard.

Mirrors the architecture of the original Operations dashboard:
- Opens the workbook (data_only=True to read cached formula results)
- Falls back to an isolated Excel snapshot copy if the workbook is locked
- Parses all KPI sections up to LIVE_DATA_MAX_ROW
- Writes dashboard_data.json with a generatedAt / sourceModifiedAt stamp

Column layout is confirmed against the actual uploaded workbook
(Operations_Data.xlsx) -- note this differs slightly from the original
handover doc in two places:
  - SKU Picked by Month has TWO branch columns: 2026 CPT (AA) and
    2026 GEORGE (AB), not just one.
  - Liseo Assembly vs Backorders sits one column to the right of what
    the original doc described: Month=AD, Assembled=AE, Backorders=AF,
    Fill Rate=AG (doc said AC:AF).

If you point this script at a DIFFERENT branch's workbook that uses a
different layout, update WORKBOOK_LAYOUT below to match -- do not
assume the columns are identical without checking, the way this file
turned out to differ from the original doc.
"""

import json
import os
import sys
import time
import shutil
import subprocess
import tempfile
from datetime import datetime, timezone

import openpyxl

# ---------------------------------------------------------------------------
# Configuration -- EDIT THESE for your environment / branch workbook
# ---------------------------------------------------------------------------

# Primary workbook path. If missing, falls back to auto-detecting a
# single .xlsx in the same folder (see find_workbook_fallback()).
WORKBOOK_PATH = os.environ.get(
    "OPS_V2_WORKBOOK_PATH",
    r"C:\Users\Dell\OneDrive\002.Operations\Dashboard\Operations-V2\Operations Data.xlsx",
)

SHEET_NAME = "DATA"
LIVE_DATA_MAX_ROW = 100  # keep parser ranges and Excel formulas extended to this row

OUTPUT_JSON_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "dashboard_data.json"
)

# Sheet block layout: (label_col, value_col, start_row, end_row)
# Columns given as 1-indexed integers (A=1, B=2, ...)
WORKBOOK_LAYOUT = {
    "housekeeping": {
        # A = WEEK label, B = 92M series, C = 12M series
        "week_col": 1,
        "series_92m_col": 2,
        "series_12m_col": 3,
        "average_row": 3,
        "start_row": 4,
        "end_row": LIVE_DATA_MAX_ROW,
    },
    "container_jhb": {"label_col": 6, "value_col": 7, "start_row": 3, "end_row": LIVE_DATA_MAX_ROW},
    "container_george": {"label_col": 9, "value_col": 10, "start_row": 3, "end_row": LIVE_DATA_MAX_ROW},
    "urgent_orders": {"label_col": 13, "value_col": 14, "start_row": 3, "end_row": LIVE_DATA_MAX_ROW},
    "dispatch_accuracy": {"label_col": 17, "value_col": 18, "start_row": 3, "end_row": LIVE_DATA_MAX_ROW},
    "sku_share": {
        "picker_col": 20,  # T
        "count_col": 21,   # U
        "share_col": 22,   # V
        "start_row": 3,
        "end_row": LIVE_DATA_MAX_ROW,
    },
    "sku_monthly": {
        "month_col": 24,     # X
        "y2024_col": 25,     # Y
        "y2025_col": 26,     # Z
        "y2026_cpt_col": 27,     # AA
        "y2026_george_col": 28,  # AB
        "start_row": 3,
        "end_row": LIVE_DATA_MAX_ROW,
    },
    "assembly_backorders": {
        "month_col": 30,      # AD
        "assembled_col": 31,  # AE
        "backorders_col": 32, # AF
        "fill_rate_col": 33,  # AG
        "start_row": 3,
        "end_row": LIVE_DATA_MAX_ROW,
    },
}

# KPI color-rule thresholds -- kept identical to the original Operations
# dashboard per current instructions. Update if this branch needs
# different cutoffs.
KPI_COLOR_RULES = {
    "housekeeping": [
        {"max": 0.50, "color": "red"},
        {"max": 0.80, "color": "orange"},
        {"color": "green"},
    ],
    "accuracy": [  # JHB / George / Dispatch
        {"max": 0.97, "color": "red", "inclusive_max": True},
        {"max": 0.99, "color": "orange"},
        {"color": "green"},
    ],
    "urgent_orders": [
        {"max": 0.05, "color": "green", "inclusive_max": True},
        {"max": 0.10, "color": "orange", "inclusive_max": True},
        {"color": "red"},
    ],
    "assembly_fill_rate": [
        {"max": 0.70, "color": "red"},
        {"max": 0.85, "color": "orange"},
        {"color": "green"},
    ],
    # SKU 2026 YTD is rank-based (1=green, 2=orange, 3=red), applied in code.
}


def find_workbook_fallback(configured_path):
    """If the configured workbook is missing, look for a single .xlsx
    in the same folder and use that instead (handles renamed files)."""
    if os.path.exists(configured_path):
        return configured_path
    folder = os.path.dirname(configured_path)
    if not os.path.isdir(folder):
        return configured_path
    candidates = [
        os.path.join(folder, f)
        for f in os.listdir(folder)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    ]
    if len(candidates) == 1:
        print(f"[refresh] Configured workbook missing, auto-detected: {candidates[0]}")
        return candidates[0]
    return configured_path


def snapshot_locked_workbook(path):
    """Use the isolated Excel COM snapshot helper (save_excel_snapshot.ps1)
    to copy a workbook that is currently open/locked in Excel. Retries on
    the classic 0x800AC472 busy error. Returns the temp copy's path."""
    tmp_dir = tempfile.mkdtemp(prefix="ops_v2_snapshot_")
    tmp_path = os.path.join(tmp_dir, "workbook_snapshot.xlsx")
    ps_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "save_excel_snapshot.ps1")

    for attempt in range(1, 6):
        result = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy", "Bypass",
                "-File", ps_script,
                "-SourcePath", path,
                "-DestPath", tmp_path,
            ],
            capture_output=True,
            text=True,
        )
        if result.returncode == 0 and os.path.exists(tmp_path):
            return tmp_path
        if "0x800AC472" in (result.stderr or "") or "busy" in (result.stderr or "").lower():
            print(f"[refresh] Excel busy (attempt {attempt}/5), retrying...")
            time.sleep(2)
            continue
        raise RuntimeError(f"Snapshot helper failed: {result.stderr}")

    raise RuntimeError("Snapshot helper failed after retries (Excel stayed busy).")


def load_worksheet(path):
    """Load the DATA sheet, falling back to a snapshot copy if the
    workbook is locked by an open Excel session."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        return wb[SHEET_NAME], path
    except PermissionError:
        print("[refresh] Workbook is locked (open in Excel). Using snapshot fallback...")
        snap_path = snapshot_locked_workbook(path)
        wb = openpyxl.load_workbook(snap_path, data_only=True)
        return wb[SHEET_NAME], snap_path


def cleanup_snapshot(used_path, original_path):
    if used_path == original_path:
        return
    folder = os.path.dirname(used_path)
    for attempt in range(1, 4):
        try:
            shutil.rmtree(folder)
            return
        except OSError:
            time.sleep(1)


def cell(ws, row, col):
    return ws.cell(row=row, column=col).value


def parse_housekeeping(ws, cfg):
    rows = []
    for r in range(cfg["start_row"], cfg["end_row"] + 1):
        week = cell(ws, r, cfg["week_col"])
        if week in (None, ""):
            continue
        rows.append({
            "week": week,
            "value_92m": cell(ws, r, cfg["series_92m_col"]),
            "value_12m": cell(ws, r, cfg["series_12m_col"]),
        })
    average = cell(ws, cfg["average_row"], cfg["series_92m_col"])
    return {"average": average, "weekly": rows}


def parse_label_value_block(ws, cfg):
    rows = []
    for r in range(cfg["start_row"], cfg["end_row"] + 1):
        label = cell(ws, r, cfg["label_col"])
        value = cell(ws, r, cfg["value_col"])
        if label in (None, ""):
            continue
        rows.append({"label": label, "value": value})
    return rows


def parse_sku_share(ws, cfg):
    rows = []
    for r in range(cfg["start_row"], cfg["end_row"] + 1):
        picker = cell(ws, r, cfg["picker_col"])
        if picker in (None, ""):
            continue
        rows.append({
            "picker": picker,
            "count": cell(ws, r, cfg["count_col"]),
            "share": cell(ws, r, cfg["share_col"]),
        })
    return rows


def parse_sku_monthly(ws, cfg):
    rows = []
    for r in range(cfg["start_row"], cfg["end_row"] + 1):
        month = cell(ws, r, cfg["month_col"])
        if month in (None, ""):
            continue
        rows.append({
            "month": month,
            "y2024": cell(ws, r, cfg["y2024_col"]),
            "y2025": cell(ws, r, cfg["y2025_col"]),
            "y2026_cpt": cell(ws, r, cfg["y2026_cpt_col"]),
            "y2026_george": cell(ws, r, cfg["y2026_george_col"]),
        })
    return rows


def parse_assembly_backorders(ws, cfg):
    rows = []
    for r in range(cfg["start_row"], cfg["end_row"] + 1):
        month = cell(ws, r, cfg["month_col"])
        if month in (None, ""):
            continue
        rows.append({
            "month": month,
            "assembled": cell(ws, r, cfg["assembled_col"]),
            "backorders": cell(ws, r, cfg["backorders_col"]),
            "fill_rate": cell(ws, r, cfg["fill_rate_col"]),
        })
    return rows


def build_dashboard_data(ws, source_path):
    data = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceModifiedAt": datetime.fromtimestamp(
            os.path.getmtime(source_path), tz=timezone.utc
        ).isoformat(),
        "kpis": {
            "housekeeping": parse_housekeeping(ws, WORKBOOK_LAYOUT["housekeeping"]),
            "container-jhb": parse_label_value_block(ws, WORKBOOK_LAYOUT["container_jhb"]),
            "container-george": parse_label_value_block(ws, WORKBOOK_LAYOUT["container_george"]),
            "urgent-orders": parse_label_value_block(ws, WORKBOOK_LAYOUT["urgent_orders"]),
            "dispatch-accuracy": parse_label_value_block(ws, WORKBOOK_LAYOUT["dispatch_accuracy"]),
            "sku-share": parse_sku_share(ws, WORKBOOK_LAYOUT["sku_share"]),
            "sku-monthly": parse_sku_monthly(ws, WORKBOOK_LAYOUT["sku_monthly"]),
            "assembly-backorders": parse_assembly_backorders(ws, WORKBOOK_LAYOUT["assembly_backorders"]),
        },
        "colorRules": KPI_COLOR_RULES,
    }
    return data


def main():
    configured_path = WORKBOOK_PATH
    resolved_path = find_workbook_fallback(configured_path)

    if not os.path.exists(resolved_path):
        print(f"[refresh] ERROR: workbook not found at {resolved_path}")
        sys.exit(1)

    ws, used_path = load_worksheet(resolved_path)
    try:
        data = build_dashboard_data(ws, resolved_path)
    finally:
        cleanup_snapshot(used_path, resolved_path)

    with open(OUTPUT_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)

    print(f"[refresh] Wrote {OUTPUT_JSON_PATH}")


if __name__ == "__main__":
    main()
